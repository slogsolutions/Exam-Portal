# registration/admin.py
import decimal
from django.utils.html import format_html
from django.urls import reverse, path
from django.contrib import admin
from django.http import HttpResponse
import csv
import openpyxl
from openpyxl.utils import get_column_letter

import gzip
import json
import io
from django.utils import timezone

from .models import CandidateProfile
from results.models import CandidateAnswer


# -------------------------
# existing CSV exporter
# -------------------------
def export_candidate_answers(modeladmin, request, queryset):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="selected_candidates_answers.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "Army Number", "Candidate Name", "Category",
        "Paper Title", "Question ID", "Question Text", "Answer", "Submitted At"
    ])

    for candidate in queryset:
        answers = CandidateAnswer.objects.filter(candidate=candidate).select_related("candidate", "paper", "question")
        for ans in answers:
            writer.writerow([
                getattr(ans.candidate, "army_no", ""),
                ans.candidate.user.get_full_name() if ans.candidate and ans.candidate.user else "",
                getattr(ans.candidate, "category", ""),
                getattr(ans.paper, "title", ""),
                getattr(ans.question, "id", ""),
                getattr(ans.question, "text", ""),
                getattr(ans, "answer", ""),
                getattr(ans, "submitted_at", ""),
            ])
    return response

export_candidate_answers.short_description = "Export selected candidates' answers to CSV"


# -------------------------
# existing Excel exporter
# -------------------------
def export_candidates_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidates"

    columns = [
        "Army No", "Rank", "Name", "Trade", "DOB", "Father Name",
        "Enrolment No", "Date of Enrolment", "Aadhar Number", "Unit",
        "Fmn (Bde)", "Fmn (Div)", "Fmn (Corps)", "Fmn (Comd)", "Training Centre",
        "District", "State",
        "Qualification", "Level of Qualification", "NSQF Level", "Skill", "QF",
        "Category", "Center", "Shift",
        "Viva Marks", "Practical Marks",
        "Created At"
    ]

    for col_num, column_title in enumerate(columns, 1):
        c = ws.cell(row=1, column=col_num)
        c.value = column_title

    for row_num, candidate in enumerate(queryset, 2):
        data = [
            candidate.army_no,
            candidate.rank,
            candidate.name,
            candidate.trade.name if getattr(candidate, "trade", None) else "",
            candidate.dob.strftime("%Y-%m-%d") if getattr(candidate, "dob", None) else "",
            candidate.father_name,
            candidate.enrolment_no,
            candidate.doe.strftime("%Y-%m-%d") if getattr(candidate, "doe", None) else "",
            candidate.aadhar_number,
            candidate.unit,
            candidate.fmn_bde,
            candidate.fmn_div,
            candidate.fmn_corps,
            candidate.fmn_comd,
            candidate.trg_centre,
            candidate.district,
            candidate.state,
            candidate.qualification.name if getattr(candidate, "qualification", None) else "",
            candidate.level_of_qualification.name if getattr(candidate, "level_of_qualification", None) else "",
            candidate.nsqf_level.name if getattr(candidate, "nsqf_level", None) else "",
            candidate.skill.name if getattr(candidate, "skill", None) else "",
            candidate.qf.name if getattr(candidate, "qf", None) else "",
            candidate.category.name if getattr(candidate, "category", None) else "",
            candidate.center.name if getattr(candidate, "center", None) else "",
            str(candidate.shift) if getattr(candidate, "shift", None) else "",
            getattr(candidate, "viva_marks", None),
            getattr(candidate, "practical_marks", None),
            candidate.created_at.strftime("%Y-%m-%d %H:%M") if getattr(candidate, "created_at", None) else "",
        ]
        for col_num, cell_value in enumerate(data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    for i, column in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="candidates.xlsx"'
    wb.save(response)
    return response

export_candidates_excel.short_description = "Export selected candidates to Excel"


# -------------------------
# NEW: binary (.dat) exporter (gzip-compressed JSON)
# -------------------------
def export_candidates_dat(modeladmin, request, queryset):
    """
    Build a structured dictionary for each candidate and compress it as gzip'd JSON.
    Returns HttpResponse with .dat binary file.
    """
    export = {
        "meta": {
            "exported_at": timezone.now().isoformat(),
            "source": "ExamPortal",
            "version": "1.0",
            "count": queryset.count()
        },
        "candidates": []
    }

    for candidate in queryset.iterator():
        cand = {
            "id": candidate.pk,
            "army_no": candidate.army_no,
            "name": candidate.name,
            "user_username": candidate.user.username if getattr(candidate, "user", None) else "",
            "user_full_name": candidate.user.get_full_name() if getattr(candidate, "user", None) else "",
            "rank": candidate.rank,
            "trade": candidate.trade.name if getattr(candidate, "trade", None) else None,
            "dob": candidate.dob.isoformat() if getattr(candidate, "dob", None) else None,
            "father_name": candidate.father_name,
            "enrolment_no": candidate.enrolment_no,
            "doe": candidate.doe.isoformat() if getattr(candidate, "doe", None) else None,
            "aadhar_number": candidate.aadhar_number,
            "unit": candidate.unit,
            "fmn_bde": candidate.fmn_bde,
            "fmn_div": candidate.fmn_div,
            "fmn_corps": candidate.fmn_corps,
            "fmn_comd": candidate.fmn_comd,
            "trg_centre": candidate.trg_centre,
            "district": candidate.district,
            "state": candidate.state,
            "qualification": candidate.qualification.name if getattr(candidate, "qualification", None) else None,
            "level_of_qualification": candidate.level_of_qualification.name if getattr(candidate, "level_of_qualification", None) else None,
            "nsqf_level": candidate.nsqf_level.name if getattr(candidate, "nsqf_level", None) else None,
            "skill": candidate.skill.name if getattr(candidate, "skill", None) else None,
            "qf": candidate.qf.name if getattr(candidate, "qf", None) else None,
            "category": candidate.category.name if getattr(candidate, "category", None) else None,
            "center": candidate.center.exam_Center if getattr(candidate, "center", None) else None,
            "shift": str(candidate.shift) if getattr(candidate, "shift", None) else None,
            "viva_marks": getattr(candidate, "viva_marks", None),
            "practical_marks": getattr(candidate, "practical_marks", None),
            "created_at": candidate.created_at.isoformat() if getattr(candidate, "created_at", None) else None,
            "photograph": candidate.photograph.url if getattr(candidate, "photograph", None) and hasattr(candidate.photograph, "url") else None,
            "answers": []
        }

        answers_qs = CandidateAnswer.objects.filter(candidate=candidate).select_related("paper", "question").order_by("paper_id", "question_id")

        for ans in answers_qs:
            ans_item = {
                "paper_id": getattr(ans.paper, "id", None),
                "paper_title": getattr(ans.paper, "title", None),
                "question_id": getattr(ans.question, "id", None),
                "question_text": getattr(ans.question, "text", None),
                "question_part": getattr(ans.question, "part", None),
                "answer": getattr(ans, "answer", None),
                "submitted_at": getattr(ans, "submitted_at", None).isoformat() if getattr(ans, "submitted_at", None) else None,
            }
            cand["answers"].append(ans_item)

        export["candidates"].append(cand)
        
    def custom_serializer(obj):
        if isinstance(obj, decimal.Decimal):
            return float(obj)
        if hasattr(obj, "isoformat"):  # dates
            return obj.isoformat()
        return str(obj)
    # dump -> gzip compress
    json_bytes = json.dumps(export, ensure_ascii=False, default=custom_serializer).encode("utf-8")

    gz_bytes = gzip.compress(json_bytes)

    ts = timezone.now().strftime("%Y%m%d%H%M%S")
    filename = f"candidates_export_{ts}.dat"

    response = HttpResponse(gz_bytes, content_type="application/octet-stream")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

export_candidates_dat.short_description = "Export selected candidates to binary (.dat gzip JSON)"


# -------------------------
# Admin: register model and add action, plus custom URL for "Export All"
# -------------------------
@admin.register(CandidateProfile)
class CandidateProfileAdmin(admin.ModelAdmin):
    list_display = ("army_no", "name", "user", "category", "viva_marks", "practical_marks")
    actions = [export_candidate_answers, export_candidates_excel, export_candidates_dat]

    def get_readonly_fields(self, request, obj=None):
        readonly = list(super().get_readonly_fields(request, obj))
        if request.user.username != "PO":
            readonly.extend(["viva_marks", "practical_marks"])
        return readonly

    def download_csv_link(self, obj):
        url = reverse("export_candidate_pdf", args=[obj.id])
        return format_html('<a class="button" href="{}">Download Answers pdf</a>', url)

    download_csv_link.short_description = "Export PDF"
    download_csv_link.allow_tags = True

    # Add a custom admin URL so you can export ALL candidates (not just selected)
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export-all-dat/', self.admin_site.admin_view(self.export_all_dat_view), name='registration_candidateprofile_export_all_dat'),
        ]
        return custom_urls + urls

    def export_all_dat_view(self, request):
        """
        Exports all candidates (respecting admin's queryset permissions).
        Accessible at: /admin/registration/candidateprofile/export-all-dat/
        """
        qs = self.get_queryset(request)
        return export_candidates_dat(self, request, qs)
